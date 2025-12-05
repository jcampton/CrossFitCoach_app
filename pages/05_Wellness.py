# 05_Wellness.py
# Wellness entry page (append-only) with:
# - Date displayed AND stored as dd/MM/yyyy (via st.date_input format=)
# - HRV Status limited to low/baseline/high
# - Sleep Score + Bed/Wake -> auto SleepQuality_1to5
# - Illness/Injury/Alcohol as Yes/No (stored 0/1)
# - BodyBattery_Start in the same row as Autonomics (compact layout)

from __future__ import annotations

import json, uuid, getpass
from datetime import datetime, timedelta, date
from pathlib import Path

import pandas as pd
from pandas.api.types import is_datetime64_any_dtype
import streamlit as st
from openpyxl import load_workbook, Workbook

# ---------------------------- Config ----------------------------

DEFAULT_TIMEZONE = "Australia/Sydney"
DEFAULT_SOURCE = "manual"
DEFAULT_ENTRY_METHOD = "ui"
SHEET_NAME = "Wellness"
DATE_UI_FORMAT = "DD/MM/YYYY"   # <-- controls the input widget display

WELLNESS_COLUMNS = [
    "WellnessID","Date","LocalTimezone",
    "Source","EntryMethod","ImportBatchID",
    "SleepStart","SleepEnd","SleepDuration_h","SleepQuality_1to5",
    "Readiness_1to100","Fatigue_1to5","Soreness_1to5","Mood_1to5",
    "IllnessFlag","InjuryFlag",
    "RestingHR_bpm","HRV_rmssd_ms","HRV_Status",
    "StressAvg_0to100","StressTotal_min",
    "BodyBattery_Start","BodyBattery_End","BodyBattery_Min","BodyBattery_Max","BodyBattery_Delta",
    "Hydration_L","Calories_kcal",
    "TravelJetlagFlag","AlcoholFlag","Notes",
    "LoggedAt_UTC","LoggedBy"
]

# ------------------------ Settings & Paths ------------------------

def load_settings(settings_path: Path) -> dict:
    try:
        with open(settings_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"data": {"workbook_path": str(Path("data") / "StrengthApp.xlsx")}}

def resolve_workbook_path(cfg: dict) -> Path:
    try:
        p = Path(cfg["data"]["workbook_path"])
    except Exception:
        p = Path("data") / "StrengthApp.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p

# ------------------------ Excel IO (append-only) ------------------------

def ensure_wellness_sheet(wb_path: Path, sheet_name: str = SHEET_NAME) -> None:
    """Create workbook/sheet + headers if missing. Never deletes/overwrites."""
    if not wb_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(WELLNESS_COLUMNS)
        wb.save(wb_path); return

    wb = load_workbook(wb_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name).append(WELLNESS_COLUMNS)
        wb.save(wb_path); return

    ws = wb[sheet_name]
    if [c.value for c in ws[1]] != WELLNESS_COLUMNS:
        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        wb.create_sheet(f"{sheet_name}_v{ts}").append(WELLNESS_COLUMNS)
        wb.save(wb_path)

def append_wellness_row(wb_path: Path, row_dict: dict, sheet_name: str = SHEET_NAME) -> None:
    wb = load_workbook(wb_path)
    ws = wb[sheet_name]
    ws.append([row_dict.get(col, "") for col in WELLNESS_COLUMNS])
    wb.save(wb_path)

def read_recent_rows(wb_path: Path, sheet_name: str = SHEET_NAME, n: int = 14) -> pd.DataFrame:
    try:
        df = pd.read_excel(wb_path, sheet_name=sheet_name, engine="openpyxl")
        # Ensure Date shows as dd/MM/yyyy in preview
        if "Date" in df.columns and is_datetime64_any_dtype(df["Date"]):
            df["Date"] = df["Date"].dt.strftime("%d/%m/%Y")
        return df.tail(n)
    except Exception:
        return pd.DataFrame(columns=WELLNESS_COLUMNS)

# ----------------------------- Helpers -----------------------------

def make_wellness_id(date_str_ddmmyyyy: str) -> str:
    return f"WELL-{date_str_ddmmyyyy}-{uuid.uuid4().hex[:4].upper()}"

def parse_datetime_local(date_val: date, hm: str | None) -> str:
    """Combine date + 'HH:MM' → 'YYYY-MM-DD HH:MM'. Return '' if blank/invalid."""
    if not date_val or not hm:
        return ""
    try:
        hh, mm = [int(x) for x in hm.split(":")]
        dt = datetime(date_val.year, date_val.month, date_val.day, hh, mm)
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return ""

def compute_sleep_duration_h(sleep_start: str, sleep_end: str) -> float | str:
    if not sleep_start or not sleep_end:
        return ""
    try:
        s = datetime.strptime(sleep_start, "%Y-%m-%d %H:%M")
        e = datetime.strptime(sleep_end, "%Y-%m-%d %H:%M")
        if e < s:
            e += timedelta(days=1)
        return round((e - s).total_seconds() / 3600.0, 2)
    except Exception:
        return ""

def safe_float(x):
    if x in ("", None):
        return ""
    try:
        return float(x)
    except Exception:
        return ""

def yesno_to_flag(val: str) -> int:
    return 1 if str(val).strip().lower() in ("yes", "y", "true", "1") else 0

def sleep_quality_from_score(score: int | None) -> tuple[int | str, str]:
    """Map Garmin Sleep Score → (SleepQuality_1to5, label)."""
    if score is None or score == "":
        return "", "Unknown"
    try:
        s = int(score)
        if s >= 90: return 5, "Excellent"
        if s >= 80: return 4, "Good"
        if s >= 60: return 3, "Fair"
        return 2, "Poor"
    except Exception:
        return "", "Unknown"

# ------------------------------- UI --------------------------------

st.title("Wellness (Append-only)")

cfg = load_settings(Path("data") / "settings.json")
wb_path = resolve_workbook_path(cfg)
st.caption(f"Workbook: {wb_path.resolve()}")

ensure_wellness_sheet(wb_path)

with st.form("wellness_form"):
    # Date widget displays dd/MM/yyyy; we'll also write that exact string to the sheet
    try:
        date_val = st.date_input("Date (dd/MM/yyyy)", value=datetime.now().date(), format=DATE_UI_FORMAT)
    except TypeError:
        # Fallback for older Streamlit versions without the 'format' kwarg
        date_val = st.date_input("Date (dd/MM/yyyy)", value=datetime.now().date())

    st.divider()
    st.subheader("Sleep")
    c1, c2, c3, c4 = st.columns([1,1,1,1])
    with c1:
        sleep_score = st.number_input("Sleep Score (0–100)", min_value=0, max_value=100, value=80, step=1)
    with c2:
        bed_time = st.text_input("Bed time (HH:MM)", placeholder="22:45")
    with c3:
        wake_time = st.text_input("Wake time (HH:MM)", placeholder="06:15")
    with c4:
        q_val, q_lbl = sleep_quality_from_score(sleep_score)
        st.metric("Quality (auto)", f"{q_lbl}", f"{q_val if q_val!='' else '-'} / 5")

    st.divider()
    st.subheader("Subjective")
    s1, s2, s3 = st.columns([1,1,1])
    with s1:
        fatigue = st.number_input("Fatigue (1–5)", min_value=1, max_value=5, value=2, step=1)
    with s2:
        soreness = st.number_input("Soreness (1–5)", min_value=1, max_value=5, value=2, step=1)
    with s3:
        mood = st.number_input("Mood (1–5)", min_value=1, max_value=5, value=4, step=1)

    s4, s5, s6 = st.columns([1,1,1])
    with s4:
        illness_yn = st.radio("Illness", ["No","Yes"], index=0, horizontal=True)
    with s5:
        injury_yn = st.radio("Injury", ["No","Yes"], index=0, horizontal=True)
    with s6:
        alcohol_yn = st.radio("Alcohol", ["No","Yes"], index=0, horizontal=True)

    st.divider()
    st.subheader("Autonomics & Body Battery")
    a1, a2, a3, a4 = st.columns([1,1,1,1])
    with a1:
        resting_hr = st.text_input("Resting HR (bpm)", placeholder="51")
    with a2:
        hrv_rmssd = st.text_input("HRV rMSSD (ms)", placeholder="62")
    with a3:
        hrv_status = st.selectbox("HRV Status", ["low","baseline","high"], index=1)  # no 'unknown'
    with a4:
        bb_start = st.text_input("BodyBattery Start (0–100)", placeholder="")

    st.divider()
    st.subheader("Notes & Nutrition")
    n1, n2 = st.columns([1,3])
    with n1:
        calories_kcal = st.text_input("Calories (kcal)", placeholder="")
    with n2:
        notes = st.text_area("Notes", value="", height=80)
    travel_yn = st.radio("Travel/Jetlag", ["No","Yes"], index=0, horizontal=True)

    submitted = st.form_submit_button("Append Wellness Row")

    if submitted:
        # Date saved as dd/MM/yyyy (string)
        date_str_ddmmyyyy = date_val.strftime("%d/%m/%Y")
        wellness_id = make_wellness_id(date_str_ddmmyyyy)

        # Sleep fields
        sleep_start = parse_datetime_local(date_val, bed_time)
        sleep_end   = parse_datetime_local(date_val, wake_time)
        sleep_dur_h = compute_sleep_duration_h(sleep_start, sleep_end)
        sleep_quality, sleep_label = sleep_quality_from_score(sleep_score)

        # Tag raw score into notes for traceability (schema unchanged)
        score_tag = f"[SleepScore={sleep_score}]"
        notes_out = (notes + " " + score_tag).strip() if score_tag not in notes else notes

        row = {
            "WellnessID": wellness_id,
            "Date": date_str_ddmmyyyy,                 # dd/MM/yyyy
            "LocalTimezone": DEFAULT_TIMEZONE,
            "Source": DEFAULT_SOURCE,
            "EntryMethod": DEFAULT_ENTRY_METHOD,
            "ImportBatchID": "",

            "SleepStart": sleep_start,
            "SleepEnd": sleep_end,
            "SleepDuration_h": sleep_dur_h,
            "SleepQuality_1to5": sleep_quality if sleep_quality != "" else "",

            "Readiness_1to100": "",
            "Fatigue_1to5": int(fatigue) if fatigue else "",
            "Soreness_1to5": int(soreness) if soreness else "",
            "Mood_1to5": int(mood) if mood else "",
            "IllnessFlag": yesno_to_flag(illness_yn),
            "InjuryFlag": yesno_to_flag(injury_yn),

            "RestingHR_bpm": safe_float(resting_hr),
            "HRV_rmssd_ms": safe_float(hrv_rmssd),
            "HRV_Status": hrv_status,

            # stress/hydration left blank by design
            "StressAvg_0to100": "",
            "StressTotal_min": "",

            # body battery (only Start kept in UI)
            "BodyBattery_Start": safe_float(bb_start),
            "BodyBattery_End": "",
            "BodyBattery_Min": "",
            "BodyBattery_Max": "",
            "BodyBattery_Delta": "",

            "Hydration_L": "",
            "Calories_kcal": safe_float(calories_kcal),
            "TravelJetlagFlag": yesno_to_flag(travel_yn),
            "AlcoholFlag": yesno_to_flag(alcohol_yn),
            "Notes": notes_out,

            "LoggedAt_UTC": datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
            "LoggedBy": getpass.getuser() or "unknown",
        }

        ensure_wellness_sheet(wb_path)
        append_wellness_row(wb_path, row)
        st.success(f"Saved Wellness for {date_str_ddmmyyyy} • Sleep {sleep_score} ({sleep_label})")
        st.caption("Append-only: existing data was not modified.")

st.divider()
st.subheader("Recent entries")
st.dataframe(read_recent_rows(wb_path, n=14), use_container_width=True)
