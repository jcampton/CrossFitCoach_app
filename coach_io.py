# coach_io.py (updated with shared load computation for REF1RM/SELF1RM/FIXED)
from __future__ import annotations

from pathlib import Path
from typing import Dict, Any, Optional, List
import json
import pandas as pd
from datetime import datetime
import uuid
import re
from math import isfinite
from pandas.api.types import is_object_dtype
from openpyxl import load_workbook, Workbook

# ──────────────────────────────────────────────────────────────────────────────
# Arrow-safe display helper (centralised for all pages)
# ──────────────────────────────────────────────────────────────────────────────

# A sensible default map for tables we render in Streamlit
ARROW_SAFE_DEFAULTS: Dict[str, str] = {
    "SessionID": "string",
    "MovementID": "string",
    "MovementName": "string",
    "Modality": "string",
    "Unit": "string",
    "Sets": "Int64",        # nullable integer
    "Reps": "string",       # keep patterns like "3(1+1)" as text
    "Load": "Float64",      # nullable float
    "%1RM": "string",
    "Notes": "string",
    "Timestamp": "string",  # parse to datetime where needed
}

def ensure_arrow_compat(
    df: pd.DataFrame,
    dtype_map: Dict[str, str] | None = None
) -> pd.DataFrame:
    """
    Return an Arrow-friendly copy of df:
      1) Apply explicit dtypes from dtype_map (overriding defaults).
      2) Convert any remaining object columns to pandas 'string'
         to avoid Arrow bytes/int confusion.
    """
    dm = {**ARROW_SAFE_DEFAULTS, **(dtype_map or {})}
    out = df.copy()

    # Apply requested dtypes where the column exists
    for col, dtype in dm.items():
        if col in out.columns:
            if dtype.startswith("datetime"):
                out[col] = pd.to_datetime(out[col], errors="coerce")
            else:
                try:
                    out[col] = out[col].astype(dtype, errors="ignore")
                except Exception:
                    # Last resort: stringify
                    out[col] = out[col].astype("string")

    # Final sweep: any lingering object → string
    for col in out.columns:
        try:
            if is_object_dtype(out[col]):
                out[col] = out[col].astype("string")
        except Exception:
            out[col] = out[col].astype("string")

    return out


# ──────────────────────────────────────────────────────────────────────────────
# Settings & paths
# ──────────────────────────────────────────────────────────────────────────────

def load_settings(settings_path: Optional[str] = None) -> Dict[str, Any]:
    cfg: Dict[str, Any] = {
        "project_root": ".",
        "data_workbook": "data/CrossFit_AI_Coach_Baseline.xlsx",
        "logs_dir": "data",
        "log_csv": "SessionMovements_Log.csv",
        "sheets": {
            "movements": "MovementLibrary",
            "programs": "Programs",
            "sessions": "Sessions",
            "session_movements": "SessionMovements",
            "one_rms": "OneRMs",
        },
        "columns": {
            "movements": {"MovementID": "MovementID", "Name": "Name", "Modality": "Modality", "Category": "Category", "Unit": "Unit"},
            "programs": {"ProgramID": "ProgramID", "ProgramName": "ProgramName", "Name": "ProgramName"},
            "sessions": {"SessionID": "SessionID", "Date": "Date", "ProgramID": "ProgramID", "Notes": "Notes"},
            "session_movements": {
                "SessionID": "SessionID",
                "MovementID": "MovementID",
                "Pct1RM": "Pct1RM",
                "Calculated Load": "Calculated Load",
                "Load": "Load",
            },
            "one_rms": {"MovementID": "MovementID", "OneRM": "OneRM"},
        },
        "display": {"round_increment": 2.5, "default_unit": "kg"},
        "data_path": "data/CrossFit_AI_Coach_Baseline.xlsx",  # alias used by 01/02
    }
    if settings_path:
        p = Path(settings_path).expanduser()
        if p.exists():
            with p.open("r", encoding="utf-8") as f:
                overrides = json.load(f)

            def deep_merge(a, b):
                if not isinstance(a, dict) or not isinstance(b, dict):
                    return b if b is not None else a
                out = dict(a)
                for k, v in b.items():
                    if isinstance(v, dict) and isinstance(out.get(k), dict):
                        out[k] = deep_merge(out[k], v)
                    elif v is not None:
                        out[k] = v
                return out

            cfg = deep_merge(cfg, overrides)
    return cfg


def derive_paths(cfg: Dict[str, Any]) -> Dict[str, Path]:
    """Return canonical paths used across pages."""
    # Maintain backward compatibility with 01/02 which reference cfg["data_path"]
    data_path = cfg.get("data_path") or cfg.get("data_workbook", "data/CrossFit_AI_Coach_Baseline.xlsx")
    base_excel = Path(data_path).expanduser().resolve()
    base_excel.parent.mkdir(parents=True, exist_ok=True)

    logs_dir = Path(cfg.get("logs_dir", base_excel.parent)).expanduser().resolve()
    logs_dir.mkdir(parents=True, exist_ok=True)

    log_csv = Path(cfg.get("log_csv", "SessionMovements_Log.csv"))
    if not log_csv.is_absolute():
        log_csv = logs_dir / log_csv

    closed_csv = base_excel.with_name("Sessions_Closed.csv")
    return {"excel": base_excel, "log_csv": log_csv, "closed_csv": closed_csv}


# ──────────────────────────────────────────────────────────────────────────────
# Excel IO helpers
# ──────────────────────────────────────────────────────────────────────────────

def load_excel_tables(workbook_path: Path) -> Dict[str, pd.DataFrame]:
    p = Path(workbook_path)
    if not p.exists():
        return {}
    xls = pd.ExcelFile(p)
    return {s: xls.parse(s) for s in xls.sheet_names}


def append_rows_to_sheet(workbook_path: Path | str, sheet_name: str, new_rows: pd.DataFrame) -> None:
    """
    Append rows to an Excel sheet using openpyxl to preserve existing formulas and formatting.
    - If the workbook/sheet doesn't exist, create it and write headers.
    - If headers are missing/new, extend headers safely before appending rows.
    - Values are appended; formulas already present in other rows are preserved by openpyxl.
    """
    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    df = (new_rows or pd.DataFrame()).copy()
    if df.empty:
        return
    df.columns = [str(c) for c in df.columns]

    if not workbook_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        # write headers
        ws.append(list(df.columns))
        # append rows
        for _, row in df.iterrows():
            ws.append([row.get(col) if isinstance(row, dict) else row[col] for col in df.columns])
        wb.save(workbook_path)
        return

    wb = load_workbook(workbook_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(list(df.columns))

    # Grab existing headers (row 1)
    existing_headers: List[str] = []
    if ws.max_row >= 1:
        existing_headers = [cell.value if cell.value is not None else "" for cell in ws[1]]

    if not existing_headers or all(h == "" for h in existing_headers):
        # write headers
        ws.delete_rows(1, ws.max_row)
        ws.append(list(df.columns))
        existing_headers = list(df.columns)

    # Union of headers (keep order: existing first, then any new)
    headers = list(existing_headers)
    for col in df.columns:
        if col not in headers:
            headers.append(col)

    # If header list changed, rewrite header row
    if headers != existing_headers:
        for j, h in enumerate(headers, start=1):
            ws.cell(row=1, column=j, value=h)

    # Append rows mapping to headers
    for _, row in df.iterrows():
        values = []
        for h in headers:
            values.append(row[h] if h in df.columns else None)
        ws.append(values)

    wb.save(workbook_path)


# ──────────────────────────────────────────────────────────────────────────────
# Logs helpers (used by 01/02)
# ──────────────────────────────────────────────────────────────────────────────

_LOG_COLS: List[str] = [
    "LogID",
    "SessionID",
    "MovementID",
    "Sets_Prescribed",
    "Reps_Prescribed",
    "Load_Prescribed",
    "Pct1RM_Prescribed",
    "Sets_Actual",
    "Reps_Actual",
    "Load_Actual",
    "RPE",
    "Notes",
    "Timestamp",
]

def ensure_log_csv(path: Path) -> None:
    p = Path(path)
    if not p.exists():
        pd.DataFrame(columns=_LOG_COLS).to_csv(p, index=False)

def append_log(path: Path, row: dict) -> None:
    ensure_log_csv(path)
    row = dict(row or {})
    row.setdefault("LogID", str(uuid.uuid4()))
    row.setdefault("Timestamp", datetime.now().isoformat(timespec="seconds"))
    out = {col: row.get(col, "") for col in _LOG_COLS}
    pd.DataFrame([out], columns=_LOG_COLS).to_csv(path, mode="a", header=False, index=False)

def read_log(path: Path) -> pd.DataFrame:
    ensure_log_csv(path)
    return pd.read_csv(path)

# Optional: typed reader (non-breaking). Use this in pages if you want stable dtypes.
LOG_DTYPE_MAP: Dict[str, str] = {
    "LogID": "string",
    "SessionID": "string",
    "MovementID": "string",
    "Sets_Prescribed": "Int64",
    "Reps_Prescribed": "Int64",
    "Load_Prescribed": "Float64",
    "Pct1RM_Prescribed": "Float64",
    "Sets_Actual": "Int64",
    "Reps_Actual": "Int64",
    "Load_Actual": "Float64",
    "RPE": "Int64",
    "Notes": "string",
    "Timestamp": "string",  # parse to dt where needed
}

def read_log_typed(path: Path, usecols: Optional[List[str]] = None) -> pd.DataFrame:
    """Read logs with consistent dtypes. Leaves datetime parsing to the caller."""
    ensure_log_csv(path)
    try:
        if usecols is None:
            dtype = {k: v for k, v in LOG_DTYPE_MAP.items() if v != "datetime64[ns]"}
        else:
            dtype = {k: v for k, v in LOG_DTYPE_MAP.items() if k in usecols and v != "datetime64[ns]"}
        df = pd.read_csv(path, usecols=usecols, dtype=dtype)
    except Exception:
        df = pd.DataFrame(columns=usecols or _LOG_COLS)
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Prescribing & parsing helpers (existing utilities)
# ──────────────────────────────────────────────────────────────────────────────

def normalize_colname(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).lower())

def pick_series(df: pd.DataFrame, candidates: list[str], default=pd.NA) -> pd.Series:
    if df.empty:
        return pd.Series([], dtype="float64")
    norm_map = {normalize_colname(c): c for c in df.columns}
    wants = [normalize_colname(c) for c in candidates]
    for w in wants:
        if w in norm_map:
            return df[norm_map[w]]
    for w in wants:
        for nn, orig in norm_map.items():
            if nn.startswith(w) or w in nn:
                return df[orig]
    return pd.Series([default] * len(df), index=df.index)

def parse_pct(value):
    if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
        return None
    try:
        s = str(value).strip()
        if s.endswith("%"):
            return float(s[:-1]) / 100.0
        f = float(s)
        return f / 100.0 if f > 1.5 else f
    except Exception:
        return None

def build_one_rm_lookup(one_rms_df: pd.DataFrame) -> dict[str, float]:
    if one_rms_df.empty:
        return {}
    candidates_mid = ["MovementID", "Movement Id", "MovID", "Movement"]
    candidates_orm = ["OneRM", "1RM", "One RM", "RM1"]
    mid = next((c for c in candidates_mid if c in one_rms_df.columns), None)
    orm = next((c for c in candidates_orm if c in one_rms_df.columns), None)
    if not mid or not orm:
        return {}
    df = one_rms_df[[mid, orm]].dropna()
    df[orm] = pd.to_numeric(df[orm], errors="coerce")
    df = df.dropna(subset=[orm])
    return dict(zip(df[mid].astype(str), df[orm].astype(float)))

def prescribe_from_pct(one_rm_lookup: Dict[str, float], movement_id: str, pct_value):
    if not movement_id or movement_id not in one_rm_lookup:
        return None
    try:
        if isinstance(pct_value, str) and pct_value.endswith("%"):
            pct = float(pct_value.strip("%")) / 100.0
        else:
            pct = float(pct_value)
            if pct > 1.5:
                pct = pct / 100.0
    except Exception:
        return None
    return float(one_rm_lookup[movement_id]) * pct

def round_to_increment(value, increment: float = 2.5):
    if value is None:
        return None
    if increment and increment > 0:
        return round(value / increment) * increment
    return value

def calc_load(one_rm_lookup: Dict[str, float], movement_id: str, pct_value, round_inc: float | int = 0.0):
    """Compute load from %1RM (fraction or percent), optionally rounded."""
    pct = parse_pct(pct_value)
    if pct is None:
        return None
    if str(movement_id) not in one_rm_lookup:
        return None
    val = one_rm_lookup[str(movement_id)] * float(pct)
    try:
        r = float(round_inc)
        if r > 0:
            return round_to_increment(val, r)
    except Exception:
        pass
    return float(val)

# ──────────────────────────────────────────────────────────────────────────────
# IDs / utilities
# ──────────────────────────────────────────────────────────────────────────────

def new_session_id() -> str:
    return f"S_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{str(uuid.uuid4().int)[:6]}"

# ──────────────────────────────────────────────────────────────────────────────
# NEW: Shared load computation for MovementLibrary REF1RM/SELF1RM/FIXED
# ──────────────────────────────────────────────────────────────────────────────

ALLOWED_LOADSOURCE = {"SELF1RM", "REF1RM", "FIXED"}

def _normalize_percent(pct) -> float | None:
    """Accept 0.70, 70, or '70%'. Return fraction or None."""
    try:
        s = str(pct).strip()
        if s.endswith("%"):
            p = float(s[:-1]) / 100.0
        else:
            p = float(s)
            if p > 1.5:
                p = p / 100.0
        return p if p >= 0 else None
    except Exception:
        return None

def _round_inc_safe(x: float, inc: float) -> float:
    try:
        inc = float(inc)
        if inc and inc > 0:
            return round(max(0.0, float(x)) / inc) * inc
        return max(0.0, float(x))
    except Exception:
        return max(0.0, float(x))

def _settings_round_inc(settings: dict) -> float:
    # support both new and legacy keys
    try:
        return float(
            settings.get("display", {}).get("round_increment",
                settings.get("default_round_increment", 2.5))
        )
    except Exception:
        return 2.5

def get_current_1rm(
    prs: pd.DataFrame | dict[str, float],
    movement_id: str,
    est_df: pd.DataFrame | None = None,
    bench_df: pd.DataFrame | None = None
) -> float:
    """
    Priority: actual PR (OneRMs sheet/dict) → estimated 1RM → benchmark target → 0
    - prs can be a DataFrame (columns: MovementID, OneRM) or a dict lookup.
    """
    # dict path
    if isinstance(prs, dict):
        try:
            val = prs.get(movement_id)
            return float(val) if val is not None and pd.notna(val) else 0.0
        except Exception:
            return 0.0

    # df path
    try:
        r = prs.loc[prs["MovementID"] == movement_id]
        if not r.empty and pd.notna(r["OneRM"].iloc[0]):
            return float(r["OneRM"].iloc[0])
    except Exception:
        pass

    if est_df is not None and not est_df.empty:
        try:
            r = est_df.loc[est_df["MovementID"] == movement_id]
            if not r.empty and "Est1RM" in r.columns and pd.notna(r["Est1RM"].iloc[0]):
                return float(r["Est1RM"].iloc[0])
        except Exception:
            pass

    if bench_df is not None and not bench_df.empty:
        try:
            r = bench_df.loc[bench_df["MovementID"] == movement_id]
            if not r.empty and "Target1RM" in r.columns and pd.notna(r["Target1RM"].iloc[0]):
                return float(r["Target1RM"].iloc[0])
        except Exception:
            pass

    return 0.0

def compute_prescribed_loads(
    session_df: pd.DataFrame,
    movement_library: pd.DataFrame,
    prs_df_or_lookup: pd.DataFrame | dict[str, float],
    settings: dict,
    est_df: pd.DataFrame | None = None,
    bench_df: pd.DataFrame | None = None
) -> pd.DataFrame:
    """
    Enriches session_df with:
      - RefMovementID (resolved reference movement for %1RM)
      - Base1RM       (float 1RM used)
      - PercentUsed   (float final % applied as fraction, e.g., 0.7)
      - CalcLoad      (unrounded)
      - FinalLoad     (rounded using per-movement or default increment)
      - LoadNote      (guard-rail note for UI)
    Rules:
      - If Load is provided and >0 → use it (Fixed kg provided)
      - Else use row %1RM if present (supports %1RM / Pct1RM / Percent)
      - Else use MovementLibrary.DefaultPercentOfRef when LoadSource is REF1RM/SELF1RM
      - Only compute for kg/lb units; others get 0 with "Unit not load-bearing"
      - If LoadSource is REF1RM, use LoadRefMovementID as the base movement for 1RM
      - If ComplexID/LoadAnchor columns exist: copy anchor FinalLoad to all rows in group
    """
    if session_df is None or session_df.empty:
        return session_df

    df = session_df.copy()

    # Prepare movement library hints
    lib = movement_library.copy()
    if "MovementID" not in lib.columns:
        raise ValueError("movement_library must include 'MovementID' column.")

    needed_lib_cols = ["MovementID", "LoadSource", "LoadRefMovementID", "DefaultPercentOfRef", "RoundingIncrement"]
    for c in needed_lib_cols:
        if c not in lib.columns:
            lib[c] = None
    lib = lib[needed_lib_cols].drop_duplicates()

    df = df.merge(lib, on="MovementID", how="left")  # adds the library columns

    default_inc = _settings_round_inc(settings)

    # Convenience accessor for per-row percent input
    def _row_pct(row) -> float | None:
        for key in ("%1RM", "Pct1RM", "Percent1RM", "Percent"):
            if key in row and pd.notna(row[key]) and str(row[key]).strip() != "":
                return _normalize_percent(row[key])
        return None

    ref_mov_list: List[str] = []
    base_1rm_list: List[float] = []
    pct_used_list: List[Optional[float]] = []
    calc_list: List[float] = []
    final_list: List[float] = []
    note_list: List[str] = []

    for _, row in df.iterrows():
        mov = row.get("MovementID")
        unit = (row.get("Unit") or "").strip().lower()

        # Only kg/lb rows are load-bearing for bar math
        if unit not in ("kg", "lb"):
            ref_mov_list.append(mov)
            base_1rm_list.append(0.0)
            pct_used_list.append(None)
            calc_list.append(0.0)
            final_list.append(0.0)
            note_list.append("Unit not load-bearing")
            continue

        # Resolve source and reference movement
        src = str(row.get("LoadSource") or "").upper()
        if src not in ALLOWED_LOADSOURCE:
            src = "SELF1RM"  # safe default

        ref_mov = mov if src == "SELF1RM" else (row.get("LoadRefMovementID") or mov)

        # Determine % precedence: row value > library default
        pct_row = _row_pct(row)
        pct_default = _normalize_percent(row.get("DefaultPercentOfRef"))
        percent = pct_row if pct_row is not None else pct_default

        # Rounding increment (movement-level beats settings)
        try:
            inc = float(row.get("RoundingIncrement")) if pd.notna(row.get("RoundingIncrement")) else default_inc
        except Exception:
            inc = default_inc

        # Base 1RM
        base_1rm = get_current_1rm(prs_df_or_lookup, ref_mov, est_df, bench_df)

        # Calculation decision tree
        fixed = row.get("Load")
        calc = 0.0
        used_pct: Optional[float] = None
        note = ""

        try:
            if fixed is not None and pd.notna(fixed) and float(fixed) > 0:
                calc = float(fixed)
                note = "Fixed kg provided"
            elif percent is not None and base_1rm > 0:
                calc = base_1rm * float(percent)
                used_pct = float(percent)
                note = f"% of {ref_mov} 1RM"
            elif src == "FIXED":
                # Caller intended fixed loads; nothing provided
                note = "FIXED source but no Load"
            else:
                note = "No 1RM or % available"
        except Exception:
            note = "Bad input; fell back to 0"

        final = _round_inc_safe(calc, inc)

        ref_mov_list.append(ref_mov)
        base_1rm_list.append(base_1rm)
        pct_used_list.append(used_pct)
        calc_list.append(calc)
        final_list.append(final)
        note_list.append(note)

    df["RefMovementID"] = ref_mov_list
    df["Base1RM"] = base_1rm_list
    df["PercentUsed"] = pct_used_list
    df["CalcLoad"] = calc_list
    df["FinalLoad"] = final_list
    df["LoadNote"] = note_list

    # Optional complexes: copy anchor FinalLoad to siblings if those columns exist
    if "ComplexID" in df.columns:
        def apply_complex(group: pd.DataFrame) -> pd.DataFrame:
            if group["ComplexID"].isna().all():
                return group
            # Determine anchor (first LoadAnchor==True; else first row)
            anchor_idx = None
            if "LoadAnchor" in group.columns:
                anchor = group.index[group["LoadAnchor"] == True].tolist()
                anchor_idx = anchor[0] if len(anchor) else None
            if anchor_idx is None:
                anchor_idx = group.index[0]

            bar = float(group.loc[anchor_idx, "FinalLoad"] or 0.0)
            if not isfinite(bar) or bar <= 0:
                return group  # keep as-is if anchor had no load

            group.loc[:, "FinalLoad"] = bar
            group.loc[:, "CalcLoad"] = bar
            group.loc[:, "LoadNote"] = "Complex bar load"
            return group

        df = df.groupby(df["ComplexID"].fillna("__NO_COMPLEX__"), group_keys=False).apply(apply_complex)

    return df
